"""
2020年5月，院教学调查结果处理
1. 将原始问卷导出成markdown；
2. 统计最好的课程投票情况。
"""
import openpyxl
from datetime import datetime

data_file = 'data/78167455_0_现代工学院本科教学情况调查_94_94.xlsx'
page_file = 'data/pages.md'


wb = openpyxl.load_workbook(data_file,read_only=True)
ws = wb['Sheet1']
pg = open(page_file,'w',encoding='utf-8',errors='ignore')

ques_cols = list('IJKLMNR')
best_cols = list('OPQ')

splitter = '<div style="page-break-after: always;"></div>\n'

if True:
    pg.write('# 现代工学院本科教学调查原始问卷\n\n')
    pg.write(f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n\n')
    pg.write(f'[TOC]\n\n')
    pg.write("> 原始数据导出自问卷星，使用Python转换为Markdown，并用Typora转换为PDF。\n")
    pg.write(splitter)

wbmap = openpyxl.load_workbook('data/课程名称映射表.xlsx')
wsmap = wbmap['Sheet1']

course_names = set()
with open('data/course-list.txt',encoding='utf-8',errors='ignore') as fp:
    for line in fp:
        line = line.strip()
        if line:
            course_names.add(line)

course_map = {}  # str->str  问卷中的名称->标准名称
if True:  # 初始化映射表
    for row in range(2,wsmap.max_row+1):
        old = wsmap.cell(row,1).value
        new = wsmap.cell(row,2).value
        if not old or not new:
            print(f"[Warning] Empty map: [{old}]->[{new}]")
        elif new not in course_names:
            print(f"[Warning] Invalid map: NewName not in list: [{new}]")
        else:
            course_map[old]=new

to_map = set()
bests = [{},{},{}]  # 分别是被列为第一、第二、第三的课程名称及其次数

def proc_best(c1,c2,c3):
    """
    维护全局状态，处理三门课的选择
    """
    for c,l in zip((c1,c2,c3),bests):
        if c in course_names:  # 标准名称
            l[c]=l.get(c,0)+1
        elif course_map.get(c) is not None:
            l[course_map[c]]=l.get(course_map[c],0)+1
        else:
            print(f"[Warning] Not in map: {c}")
            to_map.add(c)


for row in range(2,ws.max_row+1):
    num = ws.cell(row,1).value
    grade = ws.cell(row,7).value[:-1]
    major = ws.cell(row,8).value
    pg.write(f"## {num} [{grade}-{major}]\n\n")

    for c in ques_cols:
        q = ws[f"{c}{1}"].value
        a = ws[f"{c}{row}"].value
        pg.write(f"> {q}\n\n")
        pg.write(f"{a}\n\n")

    best_choices = [ws[f"{c}{row}"].value for c in best_cols]
    proc_best(*best_choices)

    pg.write("> 请按序给你你上过的课程中，感觉**最好**的三门课程名称\n\n")
    for i,s in enumerate(best_choices):
        pg.write(f"{i+1}. {s}\n\n")

    pg.write(splitter)

wb.close()
pg.close()

# 整理排序表
ranks = {}  # 课程名->选择次数 （综合（=3*第一+2*第二+第三次数）,第1，第2，第3）
for course in course_names:
    rnks = tuple(map(lambda x:x.get(course,0),bests))
    if rnks != (0,0,0):
        # 至少有人选择
        r1,r2,r3 = rnks
        ranks[course] = (
            3 * r1 + 2 * r2 + r3,
            r1,r2,r3,
        )

if to_map:
    print("[fetal] some courses has not map entry.")
    for c in to_map:
        wsmap.append((c,))
    wbmap.save('data/课程名称映射表.xlsx')
else:
    # process output
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("现代工学院本科教学调查-好评课程统计表",))
    ws.append(("课程名",'综合','第一次数','第二次数','第三次数',
               '注：综合=3*第一次数+2*第二次数+1*第三次数'))
    rank_lst = list(sorted(ranks.items(),key=lambda x:x[1],reverse=True))
    for name,rnks in rank_lst:
        ws.append((
            name,rnks[0],rnks[1],rnks[2],rnks[3]
        ))
    wb.save('data/好评课程统计.xlsx')
    wb.close()


wbmap.close()
