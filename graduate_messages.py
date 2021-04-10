"""
毕业寄语转换程序。
"""
import openpyxl
import re
import typing as tp
from datetime import datetime

filename = r'data/112592489_2_现代工学院17级——毕业寄语收集一_13_13.xlsx'
outname = filename + '_dump.md'

# constants
START_COL = 7  # openpyxl下标 从1开始
TITLE_REGEX = re.compile(r'\d+、(\D+)同学')
NULL_STR = '(空)'
SHEET_NAME = 'Sheet1'


def process_header(ws) -> tp.Tuple[tp.Dict[int, str], int]:
    """
    处理标题。
    返回：姓名->列号的映射表，以及签名所在列号。
    注意：姓名映射表的列号是相对列号（0开始），但signature是绝对。
    """
    name_map = {}
    signature_col = -1
    for c in range(START_COL, ws.max_column + 1):
        v = ws.cell(1, c).value
        if re.match(TITLE_REGEX, v):
            name = re.match(TITLE_REGEX, v).groups()[0]
            name_map[c - START_COL] = name
        else:
            print("end of columns: ", v)
            signature_col = c
    print("Number of receivers: ", len(name_map))
    return name_map, signature_col


def process(filename, outname):
    wb = openpyxl.load_workbook(filename)
    md = open(outname, 'w', encoding='utf-8', errors='ignore')
    ws = wb[SHEET_NAME]
    ws: type(wb.active)
    name_map, signat = process_header(ws)
    md.write(f'# 现代工学院2017级 毕业寄语收集\n\n')
    md.write(f'Dumped from file: `{filename}`\n')
    md.write(f'Updated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
    sig_col = tuple(ws.iter_cols(signat, signat, min_row=2, values_only=True))[0]
    print(sig_col)
    print(name_map)
    for c, v in enumerate(ws.iter_cols(
            START_COL, signat - 1, min_row=2, values_only=True)):
        name = name_map[c]
        md.write(f'\n## {c + 1}. {name}\n')
        n = 0
        for i, s in enumerate(v):
            sig = sig_col[i]
            if s != NULL_STR:
                n += 1
                md.write(f'\n### {c + 1}.{n} \n')
                md.write(f'{s}\n——{sig}\n')
    md.close()
    wb.close()


if __name__ == '__main__':
    process(filename, outname)
